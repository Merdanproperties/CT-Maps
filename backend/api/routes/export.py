from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import Optional, List
from database import get_db
from models import Property
import csv
import io
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

@router.get("/csv")
async def export_csv(
    filter_type: Optional[str] = Query(None),
    min_equity: Optional[float] = None,
    municipality: Optional[str] = None,
    property_type: Optional[str] = None,
    include_vacant: Optional[bool] = None,
    include_absentee: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """Export properties to CSV"""
    query = db.query(Property)
    
    # Apply filters
    if filter_type == "high-equity" and min_equity:
        query = query.filter(
            Property.equity_estimate.isnot(None),
            Property.equity_estimate >= min_equity
        )
    elif filter_type == "vacant":
        query = query.filter(Property.is_vacant == 1)
    elif filter_type == "absentee-owners":
        query = query.filter(Property.is_absentee == 1)
    
    if municipality:
        query = query.filter(Property.municipality.ilike(f"%{municipality}%"))
    
    if property_type:
        query = query.filter(Property.property_type == property_type)
    
    if include_vacant is not None:
        query = query.filter(Property.is_vacant == (1 if include_vacant else 0))
    
    if include_absentee is not None:
        query = query.filter(Property.is_absentee == (1 if include_absentee else 0))
    
    properties = query.limit(10000).all()  # Limit to prevent memory issues
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Parcel ID',
        'Address',
        'City',
        'Municipality',
        'Zip Code',
        'Owner Name',
        'Owner Address',
        'Owner City',
        'Owner State',
        'Owner Zip',
        'Assessed Value',
        'Land Value',
        'Building Value',
        'Property Type',
        'Land Use',
        'Lot Size (sqft)',
        'Building Area (sqft)',
        'Year Built',
        'Bedrooms',
        'Bathrooms',
        'Last Sale Date',
        'Last Sale Price',
        'Estimated Equity',
        'Is Absentee Owner',
        'Is Vacant',
        'Days Since Sale'
    ])
    
    # Write data
    for prop in properties:
        writer.writerow([
            prop.parcel_id,
            prop.address or '',
            prop.city or '',
            prop.municipality or '',
            prop.zip_code or '',
            prop.owner_name or '',
            prop.owner_address or '',
            prop.owner_city or '',
            prop.owner_state or '',
            prop.owner_zip or '',
            prop.assessed_value or '',
            prop.land_value or '',
            prop.building_value or '',
            prop.property_type or '',
            prop.land_use or '',
            prop.lot_size_sqft or '',
            prop.building_area_sqft or '',
            prop.year_built or '',
            prop.bedrooms or '',
            prop.bathrooms or '',
            prop.last_sale_date.isoformat() if prop.last_sale_date else '',
            prop.last_sale_price or '',
            prop.equity_estimate or '',
            'Yes' if prop.is_absentee == 1 else 'No',
            'Yes' if prop.is_vacant == 1 else 'No',
            prop.days_since_sale or ''
        ])
    
    output.seek(0)
    
    filename = f"ct_properties_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/json")
async def export_json(
    filter_type: Optional[str] = Query(None),
    min_equity: Optional[float] = None,
    municipality: Optional[str] = None,
    property_type: Optional[str] = None,
    include_vacant: Optional[bool] = None,
    include_absentee: Optional[bool] = None,
    limit: int = Query(1000, le=10000),
    db: Session = Depends(get_db)
):
    """Export properties to JSON"""
    query = db.query(Property)
    
    # Apply filters (same as CSV)
    if filter_type == "high-equity" and min_equity:
        query = query.filter(
            Property.equity_estimate.isnot(None),
            Property.equity_estimate >= min_equity
        )
    elif filter_type == "vacant":
        query = query.filter(Property.is_vacant == 1)
    elif filter_type == "absentee-owners":
        query = query.filter(Property.is_absentee == 1)
    
    if municipality:
        query = query.filter(Property.municipality.ilike(f"%{municipality}%"))
    
    if property_type:
        query = query.filter(Property.property_type == property_type)
    
    if include_vacant is not None:
        query = query.filter(Property.is_vacant == (1 if include_vacant else 0))
    
    if include_absentee is not None:
        query = query.filter(Property.is_absentee == (1 if include_absentee else 0))
    
    properties = query.limit(limit).all()
    
    # Convert to dict
    results = []
    for prop in properties:
        results.append({
            'parcel_id': prop.parcel_id,
            'address': prop.address,
            'city': prop.city,
            'municipality': prop.municipality,
            'zip_code': prop.zip_code,
            'owner_name': prop.owner_name,
            'owner_address': prop.owner_address,
            'owner_city': prop.owner_city,
            'owner_state': prop.owner_state,
            'owner_zip': prop.owner_zip,
            'assessed_value': prop.assessed_value,
            'land_value': prop.land_value,
            'building_value': prop.building_value,
            'property_type': prop.property_type,
            'land_use': prop.land_use,
            'lot_size_sqft': prop.lot_size_sqft,
            'building_area_sqft': prop.building_area_sqft,
            'year_built': prop.year_built,
            'bedrooms': prop.bedrooms,
            'bathrooms': prop.bathrooms,
            'last_sale_date': prop.last_sale_date.isoformat() if prop.last_sale_date else None,
            'last_sale_price': prop.last_sale_price,
            'equity_estimate': prop.equity_estimate,
            'is_absentee_owner': prop.is_absentee == 1,
            'is_vacant': prop.is_vacant == 1,
            'days_since_sale': prop.days_since_sale
        })
    
    filename = f"ct_properties_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        content=json.dumps(results, indent=2, default=str),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/excel")
async def export_excel(
    filter_type: Optional[str] = Query(None),
    min_equity: Optional[float] = None,
    municipality: Optional[str] = None,
    property_type: Optional[str] = None,
    include_vacant: Optional[bool] = None,
    include_absentee: Optional[bool] = None,
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
    min_lot_size: Optional[float] = None,
    max_lot_size: Optional[float] = None,
    db: Session = Depends(get_db)
):
    """Export properties to Excel"""
    query = db.query(Property)
    
    # Apply filters (same as CSV/JSON)
    if filter_type == "high-equity" and min_equity:
        query = query.filter(
            Property.equity_estimate.isnot(None),
            Property.equity_estimate >= min_equity
        )
    elif filter_type == "high-equity":
        query = query.filter(
            Property.equity_estimate.isnot(None),
            Property.equity_estimate >= 50000
        )
    elif filter_type == "vacant":
        query = query.filter(Property.is_vacant == 1)
    elif filter_type == "absentee-owners":
        query = query.filter(Property.is_absentee == 1)
    elif filter_type == "recently-sold":
        query = query.filter(Property.last_sale_date.isnot(None))
    elif filter_type == "low-equity":
        query = query.filter(
            Property.equity_estimate.isnot(None),
            Property.equity_estimate <= 10000
        )
    
    if municipality:
        query = query.filter(Property.municipality.ilike(f"%{municipality}%"))
    
    if property_type:
        query = query.filter(Property.property_type.ilike(f"%{property_type}%"))
    
    if min_value is not None:
        query = query.filter(Property.assessed_value >= min_value)
    if max_value is not None:
        query = query.filter(Property.assessed_value <= max_value)
    
    if min_lot_size is not None:
        query = query.filter(Property.lot_size_sqft >= min_lot_size)
    if max_lot_size is not None:
        query = query.filter(Property.lot_size_sqft <= max_lot_size)
    
    if include_vacant is not None:
        query = query.filter(Property.is_vacant == (1 if include_vacant else 0))
    
    if include_absentee is not None:
        query = query.filter(Property.is_absentee == (1 if include_absentee else 0))
    
    properties = query.limit(10000).all()  # Limit to prevent memory issues
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Properties"
    
    # Define headers
    headers = [
        'Parcel ID',
        'Address',
        'City',
        'Municipality',
        'Zip Code',
        'Owner Name',
        'Owner Address',
        'Owner City',
        'Owner State',
        'Owner Zip',
        'Assessed Value',
        'Land Value',
        'Building Value',
        'Property Type',
        'Land Use',
        'Lot Size (sqft)',
        'Building Area (sqft)',
        'Year Built',
        'Bedrooms',
        'Bathrooms',
        'Last Sale Date',
        'Last Sale Price',
        'Estimated Equity',
        'Is Absentee Owner',
        'Is Vacant',
        'Days Since Sale'
    ]
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, prop in enumerate(properties, start=2):
        ws.cell(row=row_idx, column=1, value=prop.parcel_id)
        ws.cell(row=row_idx, column=2, value=prop.address or '')
        ws.cell(row=row_idx, column=3, value=prop.city or '')
        ws.cell(row=row_idx, column=4, value=prop.municipality or '')
        ws.cell(row=row_idx, column=5, value=prop.zip_code or '')
        ws.cell(row=row_idx, column=6, value=prop.owner_name or '')
        ws.cell(row=row_idx, column=7, value=prop.owner_address or '')
        ws.cell(row=row_idx, column=8, value=prop.owner_city or '')
        ws.cell(row=row_idx, column=9, value=prop.owner_state or '')
        ws.cell(row=row_idx, column=10, value=prop.owner_zip or '')
        ws.cell(row=row_idx, column=11, value=prop.assessed_value or '')
        ws.cell(row=row_idx, column=12, value=prop.land_value or '')
        ws.cell(row=row_idx, column=13, value=prop.building_value or '')
        ws.cell(row=row_idx, column=14, value=prop.property_type or '')
        ws.cell(row=row_idx, column=15, value=prop.land_use or '')
        ws.cell(row=row_idx, column=16, value=prop.lot_size_sqft or '')
        ws.cell(row=row_idx, column=17, value=prop.building_area_sqft or '')
        ws.cell(row=row_idx, column=18, value=prop.year_built or '')
        ws.cell(row=row_idx, column=19, value=prop.bedrooms or '')
        ws.cell(row=row_idx, column=20, value=prop.bathrooms or '')
        ws.cell(row=row_idx, column=21, value=prop.last_sale_date.isoformat() if prop.last_sale_date else '')
        ws.cell(row=row_idx, column=22, value=prop.last_sale_price or '')
        ws.cell(row=row_idx, column=23, value=prop.equity_estimate or '')
        ws.cell(row=row_idx, column=24, value='Yes' if prop.is_absentee == 1 else 'No')
        ws.cell(row=row_idx, column=25, value='Yes' if prop.is_vacant == 1 else 'No')
        ws.cell(row=row_idx, column=26, value=prop.days_since_sale or '')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ct_properties_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
