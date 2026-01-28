"""
Utility functions for property ownership analysis scripts
Provides standardized Excel formatting, visualization, and data processing
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re


def detect_llc(owner_name: str) -> bool:
    """
    Detect if owner name indicates an LLC or similar business entity
    
    Args:
        owner_name: Owner name string
        
    Returns:
        True if appears to be LLC/business entity
    """
    if not owner_name or pd.isna(owner_name):
        return False
    
    owner_upper = str(owner_name).upper()
    
    # Common LLC indicators
    llc_patterns = [
        r'\bLLC\b',
        r'\bL\.L\.C\.\b',
        r'\bL\.L\.C\b',
        r'\bLIMITED\s+LIABILITY\s+COMPANY\b',
        r'\bINC\.?\b',
        r'\bINCORPORATED\b',
        r'\bCORP\.?\b',
        r'\bCORPORATION\b',
        r'\bLP\b',
        r'\bL\.P\.\b',
        r'\bLIMITED\s+PARTNERSHIP\b',
        r'\bLLP\b',
        r'\bL\.L\.P\.\b',
        r'\bPC\b',
        r'\bP\.C\.\b',
        r'\bPROFESSIONAL\s+CORPORATION\b',
        r'\bTRUST\b',
        r'\bHOLDINGS\b',
        r'\bPROPERTIES\b',
        r'\bPROPERTY\s+GROUP\b',
        r'\bREALTY\b',
        r'\bREAL\s+ESTATE\b',
    ]
    
    for pattern in llc_patterns:
        if re.search(pattern, owner_upper):
            return True
    
    return False


def format_currency(value: float) -> str:
    """Format number as currency string"""
    if pd.isna(value) or value is None:
        return ''
    return f"${value:,.2f}"


def format_number(value: float, decimals: int = 0) -> str:
    """Format number with commas"""
    if pd.isna(value) or value is None:
        return ''
    return f"{value:,.{decimals}f}"


def calculate_percentage(count: int, total: int) -> float:
    """Calculate percentage"""
    if total == 0:
        return 0.0
    return round((count / total) * 100, 2)


def export_to_excel(
    dataframes: Dict[str, pd.DataFrame],
    output_file: Path,
    title: str = "Property Analysis Report",
    auto_format: bool = True
) -> None:
    """
    Export multiple DataFrames to Excel with standardized formatting
    
    Args:
        dataframes: Dictionary of {sheet_name: DataFrame}
        output_file: Path to output Excel file
        title: Report title
        auto_format: Whether to apply automatic formatting
    """
    print(f"\nExporting to Excel: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            # Clean sheet name (Excel has 31 char limit)
            clean_sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
            
            # Write DataFrame to Excel
            df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
            
            if auto_format:
                _format_excel_sheet(writer.book, writer.sheets[clean_sheet_name], df, title)
    
    print(f"✅ Excel report saved: {output_file}")


def _format_excel_sheet(workbook: Workbook, worksheet, df: pd.DataFrame, title: str) -> None:
    """Apply formatting to Excel worksheet"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Auto-adjust column widths
    for idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        
        # Calculate max width
        max_length = max(
            df[col].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col))
        )
        
        # Set width (with limits)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Format numeric columns
        if df[col].dtype in ['float64', 'int64']:
            for row_idx in range(2, len(df) + 2):
                cell = worksheet[f"{col_letter}{row_idx}"]
                cell.number_format = '#,##0' if df[col].dtype == 'int64' else '#,##0.00'
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right")
        else:
            # Format text columns
            for row_idx in range(2, len(df) + 2):
                cell = worksheet[f"{col_letter}{row_idx}"]
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def create_summary_sheet(
    summary_data: Dict[str, any],
    sheet_name: str = "Summary"
) -> pd.DataFrame:
    """
    Create a summary DataFrame for Excel export
    
    Args:
        summary_data: Dictionary of {metric_name: value}
        sheet_name: Name for the sheet
        
    Returns:
        DataFrame with summary metrics
    """
    df = pd.DataFrame({
        'Metric': list(summary_data.keys()),
        'Value': list(summary_data.values())
    })
    return df


def get_database_session():
    """Get database session for analysis scripts"""
    import sys
    from pathlib import Path
    
    # Add parent directory to path
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    
    from database import SessionLocal
    return SessionLocal()


def load_properties_from_db(
    municipality: Optional[str] = None,
    db_session = None
) -> pd.DataFrame:
    """
    Load properties from database into DataFrame
    
    Args:
        municipality: Optional municipality filter
        db_session: Database session (creates new if None)
        
    Returns:
        DataFrame with property data
    """
    if db_session is None:
        db = get_database_session()
        should_close = True
    else:
        db = db_session
        should_close = False
    
    try:
        from models import Property
        
        query = db.query(
            Property.id,
            Property.parcel_id,
            Property.address,
            Property.municipality,
            Property.owner_name,
            Property.owner_address,
            Property.owner_city,
            Property.owner_state,
            Property.owner_zip,
            Property.assessed_value,
            Property.land_value,
            Property.building_value,
            Property.total_value,
            Property.property_type,
            Property.lot_size_sqft,
            Property.is_vacant,
            Property.is_absentee
        )
        
        if municipality:
            query = query.filter(Property.municipality == municipality)
        
        # Convert to list of dicts
        properties = []
        for prop in query.all():
            properties.append({
                'id': prop.id,
                'parcel_id': prop.parcel_id,
                'address': prop.address,
                'municipality': prop.municipality,
                'owner_name': prop.owner_name,
                'owner_address': prop.owner_address,
                'owner_city': prop.owner_city,
                'owner_state': prop.owner_state,
                'owner_zip': prop.owner_zip,
                'assessed_value': prop.assessed_value,
                'land_value': prop.land_value,
                'building_value': prop.building_value,
                'total_value': prop.total_value,
                'property_type': prop.property_type,
                'lot_size_sqft': prop.lot_size_sqft,
                'is_vacant': prop.is_vacant,
                'is_absentee': prop.is_absentee
            })
        
        df = pd.DataFrame(properties)
        return df
        
    finally:
        if should_close:
            db.close()
